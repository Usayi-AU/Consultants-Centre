from datetime import datetime
from pathlib import Path

from django.conf import settings
from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook

from reports.models import CRMActionHistory, CRMActionItem, CRMActionStatus, CRMHistoryAction


def parse_date(value):
    if value is None:
        return None
    if hasattr(value, "date"):
        return value.date()
    if isinstance(value, str):
        text = value.strip()
        for fmt in ("%d/%m/%y", "%d/%m/%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
    return None


class Command(BaseCommand):
    help = "Import CRM Client Action Items workbook into CRMActionItem model."

    def add_arguments(self, parser):
        parser.add_argument("--workbook", type=str, default=None)

    def handle(self, *args, **options):
        seed_history = not CRMActionHistory.objects.exists()
        workbook_path = options["workbook"]
        if workbook_path:
            source_path = Path(workbook_path)
        else:
            xlsx_files = sorted(Path(settings.BASE_DIR).glob("*.xlsx"))
            if not xlsx_files:
                raise CommandError("No workbook found in the project root.")
            # prefer files that mention Action_Items
            source_path = xlsx_files[0]
            for p in xlsx_files:
                if "Action_Items" in p.name or "Action Items" in p.name:
                    source_path = p
                    break

        if not source_path.exists():
            raise CommandError(f"Workbook not found: {source_path}")

        workbook = load_workbook(source_path, data_only=True)
        sheet_name = None
        for name in workbook.sheetnames:
            if "CLIENT ACTION" in name.upper() or "ACTION ITEMS" in name.upper():
                sheet_name = name
                break
        if not sheet_name:
            raise CommandError("Could not find a Client Action Items sheet in the workbook.")

        ws = workbook[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        # find header row that contains CLIENT
        header_row = None
        for idx, row in enumerate(rows[:10], start=1):
            if any(str(cell).upper().strip() == "CLIENT" for cell in (row or [])):
                header_row = idx
                break

        if not header_row:
            raise CommandError("Could not locate header row with CLIENT column.")

        start_row = header_row + 1
        created = 0
        updated = 0
        for row in ws.iter_rows(min_row=start_row, values_only=True):
            # columns based on observed workbook: B=CLIENT, C=ACTION ITEM, D=STATUS, E=OWNER, F=DATE RECEIVED, G=TARGET DATE, H=UPDATE PREV, I=UPDATE THIS
            client = row[1]
            action = row[2]
            status_text = row[3]
            owner = row[4]
            target_date = parse_date(row[6]) if len(row) > 6 else None
            update_this = row[8] if len(row) > 8 else None

            if not client or not action:
                continue

            status = CRMActionStatus.IN_PROGRESS
            if status_text and str(status_text).strip().lower().startswith("completed"):
                status = CRMActionStatus.COMPLETED

            obj, created_flag = CRMActionItem.objects.update_or_create(
                client_name=str(client).strip(),
                action_item=str(action).strip(),
                defaults={
                    "crm_owner": str(owner).strip() if owner else "",
                    "progress_update": str(update_this).strip() if update_this else "",
                    "due_date": target_date,
                    "status": status,
                },
            )
            if created_flag:
                created += 1
            else:
                updated += 1

        # Merge completed items from the dedicated completed sheet when available.
        completed_sheet = None
        for name in workbook.sheetnames:
            if "COMPLETED" in name.upper():
                completed_sheet = workbook[name]
                break

        if completed_sheet:
            for row in completed_sheet.iter_rows(min_row=1, values_only=True):
                # layout observed: B=client, C=action item, D=Done, E=owner, F=date
                client = row[1] if len(row) > 1 else None
                action = row[2] if len(row) > 2 else None
                status_text = row[3] if len(row) > 3 else None
                owner = row[4] if len(row) > 4 else None
                target_date = parse_date(row[5]) if len(row) > 5 else None

                if not client or not action:
                    continue

                if not status_text or "done" not in str(status_text).strip().lower():
                    continue

                obj, created_flag = CRMActionItem.objects.update_or_create(
                    client_name=str(client).strip(),
                    action_item=str(action).strip(),
                    defaults={
                        "crm_owner": str(owner).strip() if owner else "",
                        "progress_update": str(status_text).strip() if status_text else "Completed",
                        "due_date": target_date,
                        "status": CRMActionStatus.COMPLETED,
                    },
                )
                if created_flag:
                    created += 1
                else:
                    updated += 1

        if seed_history:
            for item in CRMActionItem.objects.all():
                CRMActionHistory.objects.create(
                    action_item=item,
                    client_name=item.client_name,
                    crm_owner=item.crm_owner,
                    action_item_text=item.action_item,
                    progress_update=item.progress_update,
                    due_date=item.due_date,
                    status=item.status,
                    action_type=CRMHistoryAction.IMPORTED,
                    actor_role="system",
                    note="Seeded from workbook import",
                )

        self.stdout.write(self.style.SUCCESS(f"Imported CRM items from {source_path.name} (created={created}, updated={updated})."))
