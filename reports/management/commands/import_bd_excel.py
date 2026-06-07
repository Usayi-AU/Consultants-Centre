from datetime import datetime
from pathlib import Path

from django.conf import settings
from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook

from reports.models import IPSReviewEntry, SelfGeneratedTarget, TenderReferralEntry
from reports.utils import default_static_workbook


def parse_bool(v):
    if v is None:
        return False
    s = str(v).strip().lower()
    return s in ("√", "x", "yes", "y", "true", "1")


def parse_date(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    s = str(v).strip()
    if not s:
        return None
    fmts = ["%d %b %Y", "%d %B %Y", "%d %m %Y", "%Y-%m-%d", "%d/%m/%Y"]
    for f in fmts:
        try:
            return datetime.strptime(s, f).date()
        except Exception:
            continue
    return None


def row_text(row, index=0):
    if row is None or len(row) <= index or row[index] is None:
        return ""
    return str(row[index]).strip()


class Command(BaseCommand):
    help = "Import Business Development Excel into DB models"

    def add_arguments(self, parser):
        parser.add_argument("path", nargs="?", default=None)

    def handle(self, *args, **options):
        path = options["path"]
        if path:
            source_path = Path(path)
            if not source_path.is_absolute():
                source_path = Path(settings.BASE_DIR) / source_path
        else:
            source_path = default_static_workbook("Intellego_IPS_Business_Development_Tracker.xlsx")
            if not source_path:
                raise CommandError("No Business Development workbook found in static/.")
        if not source_path.exists():
            raise CommandError(f"File not found: {source_path}")

        IPSReviewEntry.objects.all().delete()
        TenderReferralEntry.objects.all().delete()
        SelfGeneratedTarget.objects.all().delete()
        wb = load_workbook(source_path, read_only=True)

        # IPS Reviews
        if "IPS Reviews 2025" in wb.sheetnames:
            ws = wb["IPS Reviews 2025"]
            rows = ws.iter_rows(values_only=True)
            # find header row index
            header = None
            for r in rows:
                if r and any(cell and str(cell).strip().lower().startswith("fund name") for cell in r if cell):
                    header = [str(c).strip() if c else None for c in r]
                    break
            if header:
                for r in rows:
                    if not r or not r[0]:
                        continue
                    fund_name = str(r[0]).strip()
                    ic_crm = str(r[1]).strip() if len(r) > 1 and r[1] else ""
                    administrator = str(r[2]).strip() if len(r) > 2 and r[2] else ""
                    admin_crm = str(r[3]).strip() if len(r) > 3 and r[3] else ""
                    member_fin = parse_bool(r[4]) if len(r) > 4 else False
                    financial_review = parse_bool(r[5]) if len(r) > 5 else False
                    replacement = str(r[6]).strip() if len(r) > 6 and r[6] else ""
                    date_data = parse_date(r[7]) if len(r) > 7 else None
                    date_sent = parse_date(r[8]) if len(r) > 8 else None
                    workshop = parse_bool(r[9]) if len(r) > 9 else False
                    asset_mgr = str(r[10]).strip() if len(r) > 10 and r[10] else ""
                    status_comments = str(r[11]).strip() if len(r) > 11 and r[11] else ""
                    IPSReviewEntry.objects.update_or_create(
                        fund_name=fund_name,
                        defaults={
                            "ic_crm": ic_crm,
                            "administrator": administrator,
                            "admin_crm": admin_crm,
                            "member_fin_schedules": member_fin,
                            "financial_review": financial_review,
                            "replacement_ratios": replacement,
                            "date_data_received": date_data,
                            "date_sent_to_client": date_sent,
                            "workshop_required": workshop,
                            "asset_mgr_mandates": asset_mgr,
                            "status_comments": status_comments,
                        },
                    )
            self.stdout.write("Imported IPS Reviews")

        # Tenders, Referrals & Other
        if "Tenders, Referrals & Other" in wb.sheetnames:
            ws = wb["Tenders, Referrals & Other"]
            rows = ws.iter_rows(values_only=True)
            current_section = ""
            for r in rows:
                first = row_text(r, 0)
                if not first:
                    continue
                if first.lower().startswith("client name") or first.startswith("INTELLEGO"):
                    continue
                if len([cell for cell in r if cell not in (None, "")]) == 1:
                    current_section = first
                    continue
                client_name = first
                business_type = str(r[1]).strip() if len(r) > 1 and r[1] else ""
                contact_name = str(r[2]).strip() if len(r) > 2 and r[2] else ""
                email = str(r[3]).strip() if len(r) > 3 and r[3] else ""
                date_req = parse_date(r[4]) if len(r) > 4 else None
                date_sub = parse_date(r[5]) if len(r) > 5 else None
                status_comments = str(r[6]).strip() if len(r) > 6 and r[6] else ""
                TenderReferralEntry.objects.update_or_create(
                    client_name=client_name,
                    defaults={
                        "source_section": current_section,
                        "business_type": business_type,
                        "contact_name": contact_name,
                        "email_address": email,
                        "date_requested": date_req,
                        "date_submitted": date_sub,
                        "status_comments": status_comments,
                    },
                )
            self.stdout.write("Imported Tenders/Referrals")

        # Self-Generated Targets
        if "Self-Generated Targets" in wb.sheetnames:
            ws = wb["Self-Generated Targets"]
            rows = ws.iter_rows(values_only=True)
            current_section = ""
            for r in rows:
                first = row_text(r, 0)
                if not first:
                    continue
                if first.lower().startswith("client name") or first.startswith("INTELLEGO"):
                    continue
                if len([cell for cell in r if cell not in (None, "")]) == 1:
                    current_section = first
                    continue
                client_name = first
                administrator = str(r[1]).strip() if len(r) > 1 and r[1] else ""
                business_target = str(r[2]).strip() if len(r) > 2 and r[2] else ""
                fund_size = str(r[3]).strip() if len(r) > 3 and r[3] else ""
                contact_name = str(r[4]).strip() if len(r) > 4 and r[4] else ""
                email = str(r[5]).strip() if len(r) > 5 and r[5] else ""
                date_req = parse_date(r[6]) if len(r) > 6 else None
                date_sub = parse_date(r[7]) if len(r) > 7 else None
                comments = str(r[8]).strip() if len(r) > 8 and r[8] else ""
                SelfGeneratedTarget.objects.update_or_create(
                    client_name=client_name,
                    defaults={
                        "source_section": current_section,
                        "administrator": administrator,
                        "business_target": business_target,
                        "fund_size_usd_m": fund_size,
                        "contact_name": contact_name,
                        "email_address": email,
                        "date_requested": date_req,
                        "date_submitted": date_sub,
                        "comments": comments,
                    },
                )
            self.stdout.write("Imported Self-Generated Targets")

        self.stdout.write(self.style.SUCCESS("Import complete"))
