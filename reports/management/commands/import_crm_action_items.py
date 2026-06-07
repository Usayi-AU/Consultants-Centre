from datetime import datetime
from pathlib import Path

from django.conf import settings
from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook

from reports.models import CRMActionItem, CRMActionStatus


def normalize(value):
    if value is None:
        return ""
    return str(value).strip()


def parse_due_date(value):
    if value is None or value == "":
        return None
    if hasattr(value, "date"):
        return value.date()
    text = normalize(value)
    for fmt in ("%d/%m/%y", "%d/%m/%Y", "%Y-%m-%d", "%d-%b-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


class Command(BaseCommand):
    help = "Import Client Relations action items from an Excel workbook sheet."

    def add_arguments(self, parser):
        parser.add_argument("--workbook", type=str, default=None)
        parser.add_argument("--sheet", type=str, default="Client Relations Action Items")

    def handle(self, *args, **options):
        workbook_path = options["workbook"]
        sheet_name = options["sheet"]

        if workbook_path:
            source_path = Path(workbook_path)
        else:
            xlsx_files = sorted(Path(settings.BASE_DIR).glob("*.xlsx"))
            if not xlsx_files:
                raise CommandError("No workbook found in the project root.")
            source_path = xlsx_files[0]

        if not source_path.exists():
            raise CommandError(f"Workbook not found: {source_path}")

        workbook = load_workbook(source_path, data_only=True)
        if sheet_name not in workbook.sheetnames:
            raise CommandError(
                f"Sheet '{sheet_name}' not found. Available sheets: {', '.join(workbook.sheetnames)}"
            )

        sheet = workbook[sheet_name]

        header_row_index = None
        headers = []
        for idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=25, values_only=True), start=1):
            lowered = [normalize(col).lower() for col in row]
            if "client" in " ".join(lowered) and "owner" in " ".join(lowered):
                header_row_index = idx
                headers = lowered
                break

        if not header_row_index:
            raise CommandError("Could not detect a header row with client and owner columns.")

        def column_index(*candidates):
            for i, header in enumerate(headers):
                if any(candidate in header for candidate in candidates):
                    return i
            return None

        idx_client = column_index("client")
        idx_owner = column_index("owner", "crm")
        idx_action = column_index("action item", "action", "item")
        idx_update = column_index("update", "comment", "remark", "progress")
        idx_status = column_index("status", "progress")
        idx_due = column_index("due", "target", "deadline")

        if idx_client is None or idx_owner is None or idx_action is None:
            raise CommandError("Workbook needs at least client, owner, and action item columns.")

        CRMActionItem.objects.all().delete()
        created = 0

        for row in sheet.iter_rows(min_row=header_row_index + 1, values_only=True):
            client_name = normalize(row[idx_client] if idx_client is not None else "")
            owner = normalize(row[idx_owner] if idx_owner is not None else "")
            action_item = normalize(row[idx_action] if idx_action is not None else "")
            if not client_name or not owner or not action_item:
                continue

            update_text = normalize(row[idx_update] if idx_update is not None else "")
            status_text = normalize(row[idx_status] if idx_status is not None else "").lower()
            due_date = parse_due_date(row[idx_due] if idx_due is not None else None)

            status = CRMActionStatus.IN_PROGRESS
            if "complete" in status_text or "done" in status_text:
                status = CRMActionStatus.COMPLETED

            CRMActionItem.objects.create(
                client_name=client_name,
                crm_owner=owner,
                action_item=action_item,
                progress_update=update_text,
                due_date=due_date,
                status=status,
            )
            created += 1

        self.stdout.write(self.style.SUCCESS(f"Imported {created} CRM action items from {source_path.name}:{sheet_name}."))
