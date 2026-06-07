from datetime import datetime
from pathlib import Path

from django.conf import settings
from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook

from reports.models import ClientReport, StatusPhase
from reports.utils import default_static_workbook


def parse_due_date(value):
    if value is None:
        return None
    if hasattr(value, "date"):
        return value.date()
    if isinstance(value, str):
        text = value.strip()
        for fmt in ("%d/%m/%y", "%d/%m/%Y"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
    return None


class Command(BaseCommand):
    help = "Import the Q1 2026 report tracker workbook into the database."

    def add_arguments(self, parser):
        parser.add_argument("--workbook", type=str, default=None)

    def handle(self, *args, **options):
        workbook_path = options["workbook"]
        if workbook_path:
            source_path = Path(workbook_path)
        else:
            source_path = default_static_workbook("Q1*Report Tracker*.xlsx")
            if not source_path:
                xlsx_files = sorted(Path(settings.BASE_DIR).glob("*.xlsx"))
                if not xlsx_files:
                    raise CommandError("No operations tracker workbook found in static/ or project root.")
                source_path = xlsx_files[0]

        if not source_path.exists():
            raise CommandError(f"Workbook not found: {source_path}")

        workbook = load_workbook(source_path, data_only=True)
        if "Report Tracker" not in workbook.sheetnames:
            raise CommandError("The workbook does not contain a 'Report Tracker' sheet.")

        worksheet = workbook["Report Tracker"]
        ClientReport.objects.all().delete()

        rows_created = 0
        for row in worksheet.iter_rows(min_row=5, values_only=True):
            client_name = row[0]
            crm_name = row[1]
            operations_assignee = row[2]
            submitted_mark = row[3]
            reviewed_mark = row[4]
            sent_mark = row[5]
            due_date = parse_due_date(row[6])
            status_text = row[7]

            if not client_name or str(client_name).startswith("STATUS:"):
                continue

            if sent_mark == "✓" or status_text == "Sent to Client":
                status_phase = StatusPhase.SENT_TO_CLIENT
            elif reviewed_mark == "✓" or status_text == "Reviewed":
                status_phase = StatusPhase.REVIEWED
            elif submitted_mark == "✓" or status_text == "Submitted":
                status_phase = StatusPhase.SUBMITTED
            else:
                status_phase = StatusPhase.PENDING

            ClientReport.objects.create(
                client_name=str(client_name).strip(),
                crm_name=str(crm_name).strip(),
                operations_assignee=str(operations_assignee).strip(),
                due_date=due_date,
                status_phase=status_phase,
                notes="",
            )
            rows_created += 1

        self.stdout.write(self.style.SUCCESS(f"Imported {rows_created} report rows from {source_path.name}."))
