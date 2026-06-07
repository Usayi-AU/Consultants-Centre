import re
from datetime import datetime
from pathlib import Path

import openpyxl
from django.conf import settings
from django.core.management.base import BaseCommand, CommandError
from django.utils.dateparse import parse_date

from crm.models import ActionItem, SummaryMetric, STATUS_DONE
from reports.utils import default_static_workbook


def normalize_status(value):
    if not value:
        return 'Open'
    text = str(value).strip()
    if re.search(r'done|completed|complete', text, re.I):
        return STATUS_DONE
    if re.search(r'progress|ongoing|in progress', text, re.I):
        return 'In Progress'
    if re.search(r'hold', text, re.I):
        return 'On Hold'
    return text


def parse_date_safe(value):
    if isinstance(value, datetime):
        return value.date()
    if value is None:
        return None
    return parse_date(str(value))


class Command(BaseCommand):
    help = 'Import Excel workbook data into the dashboard models'

    def add_arguments(self, parser):
        parser.add_argument('--path', default=None, help='Excel workbook path (defaults to latest Action Items file in static/)')

    def handle(self, *args, **options):
        path = options['path']
        if path:
            source_path = Path(path)
            if not source_path.is_absolute():
                source_path = Path(settings.BASE_DIR) / source_path
        else:
            source_path = default_static_workbook('Action_Items_Dashboard*.xlsx')
            if not source_path:
                raise CommandError('No Client Relations workbook found in static/.')
        if not source_path.exists():
            raise CommandError(f'Workbook not found: {source_path}')

        ActionItem.objects.all().delete()
        SummaryMetric.objects.all().delete()
        wb = openpyxl.load_workbook(source_path, data_only=True)

        if '📊 SUMMARY' in wb.sheetnames:
            ws = wb['📊 SUMMARY']
            imported_summary = False
            summary_rows = list(ws.iter_rows(min_row=5, max_row=12, values_only=True))
            for index, row in enumerate(summary_rows):
                labels = [
                    str(cell).strip()
                    for cell in (row[1:6] if row else [])
                    if isinstance(cell, str) and str(cell).strip()
                ]
                if len(labels) < 2 or index + 1 >= len(summary_rows):
                    continue
                value_row = summary_rows[index + 1]
                values = [cell for cell in (value_row[1:6] if value_row else []) if cell is not None]
                if len(values) != len(labels):
                    continue
                for label, value in zip(labels, values):
                    SummaryMetric.objects.update_or_create(
                        label=label.replace('\n', ' '),
                        defaults={
                            'value': str(value).strip() if value is not None else '',
                            'details': '',
                        },
                    )
                imported_summary = True
                break

            if not imported_summary:
                for row in ws.iter_rows(min_row=5, values_only=True):
                    if not any(row):
                        continue
                    label = row[0]
                    if not label or not isinstance(label, str):
                        continue
                    value = row[1] if len(row) > 1 else ''
                    SummaryMetric.objects.update_or_create(
                        label=label.strip(),
                        defaults={
                            'value': str(value).strip() if value is not None else '',
                            'details': '',
                        },
                    )
                    imported_summary = True
            if imported_summary:
                self.stdout.write(self.style.SUCCESS('Imported summary metrics.'))

        if '📋 CLIENT ACTION ITEMS' in wb.sheetnames:
            ws = wb['📋 CLIENT ACTION ITEMS']
            header_row = None
            for row in ws.iter_rows(values_only=True):
                if row and row[1] == 'CLIENT NAME':
                    header_row = [cell for cell in row]
                    break
            if header_row:
                start = False
                current_client = None
                for row in ws.iter_rows(values_only=True):
                    if row and row[1] == 'CLIENT NAME':
                        start = True
                        continue
                    if not start:
                        continue
                    if not any(row):
                        continue
                    client = row[1] or ''
                    # If client is empty, use the previously seen client (Excel grouping pattern)
                    if not client.strip():
                        client = current_client or ''
                    else:
                        current_client = client.strip()
                    action_item = row[2] or ''
                    status = normalize_status(row[3])
                    owner = row[4] or ''
                    date_received = parse_date_safe(row[5])
                    target_date = parse_date_safe(row[6])
                    update_prev_week = row[7] or ''
                    update_this_week = row[8] or ''
                    ActionItem.objects.update_or_create(
                        client=client.strip(),
                        action_item=action_item.strip(),
                        defaults={
                            'status': status,
                            'owner': owner.strip(),
                            'date_received': date_received,
                            'target_date': target_date,
                            'update_prev_week': str(update_prev_week).strip(),
                            'update_this_week': str(update_this_week).strip(),
                        }
                    )
                self.stdout.write(self.style.SUCCESS('Imported client action items.'))

        if '✅ COMPLETED ITEMS' in wb.sheetnames:
            ws = wb['✅ COMPLETED ITEMS']
            header_row = None
            for row in ws.iter_rows(values_only=True):
                if row and row[1] == 'CLIENT':
                    header_row = [cell for cell in row]
                    break
            if header_row:
                start = False
                current_client = None
                for row in ws.iter_rows(values_only=True):
                    if row and row[1] == 'CLIENT':
                        start = True
                        continue
                    if not start:
                        continue
                    if not any(row):
                        continue
                    client = row[1] or ''
                    # If client is empty, use the previously seen client (Excel grouping pattern)
                    if not client.strip():
                        client = current_client or ''
                    else:
                        current_client = client.strip()
                    action_item = row[2] or ''
                    status = normalize_status(row[3])
                    owner = row[4] or ''
                    completion_date = parse_date_safe(row[5])
                    ActionItem.objects.update_or_create(
                        client=client.strip(),
                        action_item=action_item.strip(),
                        defaults={
                            'status': status,
                            'owner': owner.strip(),
                            'completion_date': completion_date,
                        }
                    )
                self.stdout.write(self.style.SUCCESS('Imported completed items.'))
